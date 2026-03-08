import streamlit as st
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random
import time
import re
import io
from datetime import datetime

st.set_page_config(page_title="LinkedIn Scraper", page_icon="🔍", layout="centered")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap');
* { font-family: 'DM Sans', sans-serif; }
h1,h2,h3 { font-family: 'Syne', sans-serif !important; }
.stApp { background: #0a0a0f; color: #e8e8f0; }
.main-title { font-family:'Syne',sans-serif; font-size:2.4rem; font-weight:800; background:linear-gradient(135deg,#00c6ff,#7b2ff7,#ff6b6b); -webkit-background-clip:text; -webkit-text-fill-color:transparent; margin-bottom:0.2rem; }
.sub-title { color:#888; font-size:0.95rem; margin-bottom:2rem; font-weight:300; }
.card { background:#13131a; border:1px solid #222230; border-radius:16px; padding:1.5rem; margin-bottom:1.2rem; }
.card-title { font-family:'Syne',sans-serif; font-size:0.75rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; color:#7b2ff7; margin-bottom:1rem; }
.result-item { background:#1a1a24; border:1px solid #2a2a3a; border-radius:10px; padding:0.85rem 1rem; margin-bottom:0.6rem; }
.result-name { font-weight:500; color:#e8e8f0; font-size:0.95rem; }
.result-url { color:#00c6ff; font-size:0.78rem; margin-top:0.15rem; }
.stat-box { background:#13131a; border:1px solid #222230; border-radius:12px; padding:1rem; text-align:center; }
.stat-num { font-family:'Syne',sans-serif; font-size:2rem; font-weight:800; color:#00c6ff; }
.stat-label { color:#555; font-size:0.78rem; margin-top:0.2rem; }
div[data-testid="stButton"] button { background:linear-gradient(135deg,#7b2ff7,#00c6ff) !important; color:white !important; border:none !important; border-radius:10px !important; font-family:'Syne',sans-serif !important; font-weight:700 !important; }
div[data-testid="stDownloadButton"] button { background:#1a2a1a !important; color:#4cff91 !important; border:1px solid #4cff9140 !important; border-radius:10px !important; font-family:'Syne',sans-serif !important; font-weight:700 !important; width:100% !important; }
.stTextInput input,.stNumberInput input,.stTextArea textarea { background:#1a1a24 !important; border:1px solid #2a2a3a !important; border-radius:10px !important; color:#e8e8f0 !important; }
label { color:#aaa !important; font-size:0.85rem !important; }
.tip { background:#0d1a2a; border-left:3px solid #00c6ff; border-radius:0 8px 8px 0; padding:0.75rem 1rem; color:#7ab8d4; font-size:0.82rem; margin-top:0.5rem; }
</style>
""", unsafe_allow_html=True)

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
]

def get_headers():
    return {"User-Agent": random.choice(USER_AGENTS), "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8", "Accept-Language": "en-US,en;q=0.5", "DNT": "1"}

def extract_name(title):
    for sep in [" - ", " | ", " – ", " — "]:
        if sep in title:
            return title.split(sep)[0].strip()
    return re.sub(r"\s*[-|–—]\s*LinkedIn.*", "", title, flags=re.IGNORECASE).strip() or "Unknown"

def scrape_google(query, max_results, progress_cb=None):
    profiles, seen = [], set()
    start = 0
    while len(profiles) < max_results:
        url = f"https://www.google.com/search?q={requests.utils.quote(query)}&start={start}&num=10&hl=en"
        try:
            resp = requests.get(url, headers=get_headers(), timeout=15)
        except Exception as e:
            return profiles, f"Network error: {e}"
        if "captcha" in resp.text.lower() or "detected unusual traffic" in resp.text.lower():
            return profiles, "Google CAPTCHA triggered. Wait 10 min and try again."
        soup = BeautifulSoup(resp.text, "html.parser")
        blocks = soup.select("div.g, div.tF2Cxc, div.hlcw0c")
        found = 0
        for block in blocks:
            a = block.select_one("a[href]")
            if not a: continue
            href = a.get("href", "")
            if href.startswith("/url?"):
                m = re.search(r"[?&]q=([^&]+)", href)
                href = requests.utils.unquote(m.group(1)) if m else ""
            if "linkedin.com/in/" not in href: continue
            clean = href.split("?")[0].rstrip("/")
            if clean in seen: continue
            seen.add(clean)
            h3 = block.select_one("h3")
            name = extract_name(h3.get_text(strip=True)) if h3 else "Unknown"
            snip = block.select_one("div.VwiC3b, span.aCOpRe")
            snippet = snip.get_text(strip=True)[:200] if snip else ""
            profiles.append({"name": name, "url": clean, "snippet": snippet})
            found += 1
            if progress_cb: progress_cb(len(profiles), max_results)
            if len(profiles) >= max_results: break
        if found == 0: break
        start += 10
        time.sleep(random.uniform(2, 5))
    return profiles, None

def load_existing_urls(ws):
    urls = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        val = row[2] if len(row) >= 3 else None
        if val: urls.add(str(val).strip().rstrip("/"))
    return urls

def find_next_row(ws):
    row = 3
    while ws.cell(row=row, column=1).value:
        row += 1
    return row

def append_profiles(excel_bytes, profiles):
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb["1 - Outreach Tracker"]
    existing = load_existing_urls(ws)
    added, skipped = 0, 0
    thin = Side(style="thin", color="D0D0D0")
    border = Border(bottom=thin)
    for p in profiles:
        clean = p["url"].rstrip("/")
        if clean in existing:
            skipped += 1
            continue
        row = find_next_row(ws)
        fill = PatternFill("solid", start_color="F0F7FF" if row % 2 == 0 else "FFFFFF", end_color="F0F7FF" if row % 2 == 0 else "FFFFFF")
        for col, val, link in [(1,p["name"],False),(2,"LinkedIn",False),(3,p["url"],True),(4,"",False),(8,"Not Contacted",False)]:
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill; c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if link:
                c.hyperlink = val
                c.font = Font(color="0563C1", underline="single")
        existing.add(clean)
        added += 1
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue(), added, skipped

# ── UI ──
st.markdown('<div class="main-title">LinkedIn Scraper</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Search Google → Extract profiles → Save to your Outreach Tracker</div>', unsafe_allow_html=True)

st.markdown('<div class="card"><div class="card-title">① Upload Your Tracker</div>', unsafe_allow_html=True)
uploaded = st.file_uploader("Drop your Excel file here", type=["xlsx"], label_visibility="collapsed")
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="card"><div class="card-title">② Define Your Search</div>', unsafe_allow_html=True)
query = st.text_input("Search query", value='site:linkedin.com/in "marketing manager" India', label_visibility="collapsed")
st.markdown('<div class="tip">💡 Try: <code>site:linkedin.com/in "salon owner" India</code> &nbsp;|&nbsp; <code>site:linkedin.com/in "startup founder" Hyderabad</code></div>', unsafe_allow_html=True)
max_results = st.slider("Max profiles to scrape", 5, 50, 10)
st.markdown('</div>', unsafe_allow_html=True)

run = st.button("🔍  Scrape LinkedIn Profiles", use_container_width=True)

if run:
    if not uploaded:
        st.error("Please upload your Excel tracker first.")
    elif not query.strip():
        st.error("Please enter a search query.")
    else:
        excel_bytes = uploaded.read()
        st.markdown('<div class="card"><div class="card-title">⟳ Scraping in progress</div>', unsafe_allow_html=True)
        prog = st.progress(0)
        status_text = st.empty()
        def update_progress(current, total):
            prog.progress(min(int(current/total*100), 100))
            status_text.markdown(f'<span style="color:#888;font-size:0.85rem">Found {current} of {total} profiles...</span>', unsafe_allow_html=True)
        profiles, error = scrape_google(query, max_results, update_progress)
        st.markdown('</div>', unsafe_allow_html=True)

        if error:
            st.error(f"⚠️ {error}")
        elif not profiles:
            st.warning("No LinkedIn profiles found. Try a different query.")
        else:
            prog.progress(100)
            updated_excel, added, skipped = append_profiles(excel_bytes, profiles)
            c1,c2,c3 = st.columns(3)
            with c1: st.markdown(f'<div class="stat-box"><div class="stat-num">{len(profiles)}</div><div class="stat-label">Profiles Found</div></div>', unsafe_allow_html=True)
            with c2: st.markdown(f'<div class="stat-box"><div class="stat-num" style="color:#4cff91">{added}</div><div class="stat-label">Added to Sheet</div></div>', unsafe_allow_html=True)
            with c3: st.markdown(f'<div class="stat-box"><div class="stat-num" style="color:#555">{skipped}</div><div class="stat-label">Duplicates Skipped</div></div>', unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<div class="card"><div class="card-title">Results Preview</div>', unsafe_allow_html=True)
            for p in profiles:
                st.markdown(f'<div class="result-item"><div class="result-name">{p["name"]}</div><div class="result-url">{p["url"]}</div></div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
            fname = uploaded.name.replace(".xlsx", f"_updated_{datetime.now().strftime('%d%b')}.xlsx")
            st.download_button(label="⬇️  Download Updated Outreach Tracker", data=updated_excel, file_name=fname, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
